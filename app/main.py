# app/ingest.py
import os
import json
import re
import unicodedata
from typing import List, Tuple, Any

import pandas as pd
from bs4 import BeautifulSoup
from markdownify import markdownify as md
from docx import Document as DocxDocument
from pypdf import PdfReader

from openpyxl import load_workbook           # <- pour lire le surlignage
from openpyxl.utils import get_column_letter  # <- pour convertir index -> lettre de colonne

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import Chroma

from .config import (
    CHROMA_DIR,
    EMBEDDINGS_MODEL,
    MAX_CHUNK_TOKENS,
    CHUNK_OVERLAP_TOKENS,
    COLLECTION_NAME,
    STORE_DIR,
)

# ---------- Loaders texte classiques ----------
def _load_pdf_pages(path: str) -> List[Tuple[str, int]]:
    out: List[Tuple[str, int]] = []
    try:
        reader = PdfReader(path)
        for i, p in enumerate(reader.pages, start=1):
            out.append(((p.extract_text() or "").strip(), i))
    except Exception:
        pass
    return out

def _load_docx(path: str) -> str:
    try:
        doc = DocxDocument(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""

def _load_html(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()
    soup = BeautifulSoup(html, "lxml")
    for t in soup(["script", "style", "noscript"]):
        t.decompose()
    return md(str(soup))

def _load_txt(path: str) -> str:
    return open(path, "r", encoding="utf-8", errors="ignore").read()

# ---------- Splitter ----------
splitter = RecursiveCharacterTextSplitter(
    chunk_size=MAX_CHUNK_TOKENS,
    chunk_overlap=CHUNK_OVERLAP_TOKENS,
    separators=["\n\n", "\n", ". ", " "],
)

# ---------- Vector store ----------
def _get_vs() -> Chroma:
    return Chroma(
        persist_directory=CHROMA_DIR,
        collection_name=COLLECTION_NAME,
        embedding_function=OpenAIEmbeddings(model=EMBEDDINGS_MODEL),
    )

# ---------- Ingestion dossier (txt/pdf/docx/html) ----------
def ingest_folder(root: str, source_type: str = "mixed"):
    texts, metas = [], []
    file_count = 0

    for dirpath, _, filenames in os.walk(root):
        for fname in filenames:
            path = os.path.join(dirpath, fname)
            ext = fname.lower().rsplit(".", 1)[-1]
            title = os.path.splitext(fname)[0]

            if ext == "pdf":
                pages = _load_pdf_pages(path)
                any_page = False
                for page_text, page_no in pages:
                    if not page_text.strip():
                        continue
                    any_page = True
                    chunks = [c for c in splitter.split_text(page_text) if c.strip()]
                    texts.extend(chunks)
                    metas.extend(
                        [{"source": path, "title": title, "source_type": source_type, "page": page_no}]
                        * len(chunks)
                    )
                if any_page:
                    file_count += 1
                continue

            if ext == "docx":
                text = _load_docx(path)
            elif ext in ("html", "htm"):
                text = _load_html(path)
            elif ext in ("txt", "md"):
                text = _load_txt(path)
            else:
                continue

            if not text.strip():
                continue

            chunks = [c for c in splitter.split_text(text) if c.strip()]
            texts.extend(chunks)
            metas.extend(
                [{"source": path, "title": title, "source_type": source_type}] * len(chunks)
            )
            file_count += 1

    if not texts:
        return {"indexed_files": 0, "indexed_chunks": 0}

    vs = _get_vs()
    vs.add_texts(texts=texts, metadatas=metas)
    vs.persist()
    return {"indexed_files": file_count, "indexed_chunks": len(texts)}

# ---------- Helpers ----------
def _norm_name(s: Any) -> str:
    """Normalise un nom de colonne: lower, trim, retire accents & NBSP, compresse espaces."""
    s = str(s).replace("\u00a0", " ")  # NBSP -> espace normal
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = " ".join(s.split())
    return s.strip().lower()

def _col_lookup(df: pd.DataFrame, *aliases) -> str:
    """Retourne le nom de colonne réel (case-sensitive) à partir d'aliases tolérants."""
    low = {_norm_name(c): c for c in df.columns}
    for a in aliases:
        key = _norm_name(a)
        if key in low:
            return low[key]
    return ""

# --- Détection de sous-question & de GEN ---
_GEN_WORD = re.compile(r"\bgen\s*([123])\b", re.IGNORECASE)

def _extract_followup_question(answer_text: str) -> str:
    """
    Renvoie la première ligne courte qui finit par '?'
    (typiquement la sous-question en tête de 'Antwoord').
    """
    if not answer_text:
        return ""
    for line in answer_text.strip().splitlines()[:3]:
        s = line.strip()
        if s.endswith("?") and 3 <= len(s) <= 200:
            return s
    return ""

def _highlight_mask_for_answer(path: str, sheet_name: str | None, answer_col_letter: str) -> set[int]:
    """
    Renvoie les indices 0-based (par rapport à pandas, header exclu)
    des lignes dont la cellule 'Antwoord' a un fond non blanc.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    highlight_rows = set()
    col_idx = ord(answer_col_letter.upper()) - ord("A")  # 0-based
    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        cell = row[col_idx]
        fill = cell.fill
        rgb = getattr(getattr(fill, "start_color", None), "rgb", None)
        # Surlignage si on a une couleur non transparente et non blanche
        if fill and fill.fill_type and rgb and rgb not in ("00000000", "FFFFFFFF"):
            highlight_rows.add(i)
    return highlight_rows

def _boolish(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "x", "✓", "yes", "ja", "oui"}

# ---------- Ingestion Excel (FAQ) ----------
def ingest_excel(path: str, source_type: str = "faq"):
    """
    Lis l'Excel et :
      - choisit la feuille contenant 'Vraag'/'Antwoord',
      - écrit store/faq_index.json,
      - envoie les lignes dans Chroma (métadonnées scalaires).
    Colonnes acceptées :
      * 'Categorie'
      * 'Vraag' / 'Question'
      * 'Antwoord' / 'Answer'
      * 'Foto'
      * 'Filmpje' / 'Video' / 'Video_URL'
      * 'AskGen' (optionnel)
      * 'Gen 1' / 'Gen1' / 'Gen 2' / 'Gen2' / 'Gen 3' / 'Gen3'
      * autres colonnes cochées 'x' => tags
    """
    if not os.path.exists(path):
        return {"indexed_files": 0, "indexed_chunks": 0, "error": "file not found"}

    # 1) Choix de la feuille
    xls = pd.ExcelFile(path, engine="openpyxl")
    chosen_df, chosen_sheet = None, None
    for sheet in xls.sheet_names:
        df_try = xls.parse(sheet).fillna("")
        c_q_try = _col_lookup(df_try, "Vraag", "Question", "Vragen")
        c_a_try = _col_lookup(df_try, "Antwoord", "Answer")
        if c_q_try and c_a_try:
            chosen_df, chosen_sheet = df_try, sheet
            break

    if chosen_df is None:
        return {"indexed_files": 0, "indexed_chunks": 0, "error": "kolommen 'Vraag'/'Antwoord' ontbreken"}

    df = chosen_df

    # 2) Résolution colonnes
    c_q     = _col_lookup(df, "Vraag", "Question", "Vragen")
    c_a     = _col_lookup(df, "Antwoord", "Answer")
    c_cat   = _col_lookup(df, "Categorie", "Category")
    c_photo = _col_lookup(df, "Foto", "Photo")
    c_video = _col_lookup(df, "Filmpje", "Video", "Video_URL", "Video Url")
    c_ask   = _col_lookup(df, "AskGen", "Ask Gen", "Ask_Gen")  # optionnel

    # 3) Colonnes Gen (toutes celles dont le nom normalisé contient 'gen')
    gen_cols_real = [c for c in df.columns if "gen" in _norm_name(c)]

    # 4) Masque de surlignage sur la colonne 'Antwoord'
    try:
        a_pos = list(df.columns).index(c_a)  # position 0-based dans pandas
        a_letter = get_column_letter(a_pos + 1)  # -> 'C' dans tes fichiers
    except Exception:
        a_letter = "C"  # fallback
    highlighted = _highlight_mask_for_answer(path, chosen_sheet, a_letter)

    texts, metas, index_rows = [], [], []

    # 5) Itération lignes
    for ridx, row in enumerate(df.itertuples(index=False), start=0):
        vraag = str(getattr(row, c_q, "")).strip()
        antw  = str(getattr(row, c_a, "")).strip()
        if not vraag or not antw:
            continue

        category = str(getattr(row, c_cat, "")).strip() if c_cat else ""
        photo    = str(getattr(row, c_photo, "")).strip() if c_photo else ""
        video    = str(getattr(row, c_video, "")).strip() if c_video else ""

        # Gens (liste)
        gens_list: List[str] = []
        for gc in gen_cols_real:
            val = getattr(row, gc, "")
            if _boolish(val):
                l = _norm_name(gc)
                if "gen 1" in l or l == "gen1":
                    gens_list.append("gen1")
                elif "gen 2" in l or l == "gen2":
                    gens_list.append("gen2")
                elif "gen 3" in l or l == "gen3":
                    gens_list.append("gen3")

        # Tags = colonnes cochées 'x' hors bases & hors gen*
        base_cols = {c_q, c_a, c_cat, c_photo, c_video, c_ask}
        tags_list: List[str] = []
        for col in df.columns:
            cname = str(col)
            if cname in base_cols:
                continue
            if "gen" in _norm_name(cname):
                continue
            v = getattr(row, col, "")
            if _boolish(v):
                tags_list.append(cname)

        # --- Détection du besoin de clarification GEN ---
        followup_q = _extract_followup_question(antw)
        mentions_gen = bool(_GEN_WORD.search(followup_q))
        ask_explicit = bool(c_ask and _boolish(getattr(row, c_ask, "")))
        is_hl = ridx in highlighted

        # Règle : on demande GEN seulement si AskGen explicite
        #         OU si la sous-question mentionne GEN (avec/sans surlignage)
        #         OU si c'est surligné et qu'au moins une GEN est cochée
        ask_gen = ask_explicit or mentions_gen or (is_hl and (mentions_gen or len(gens_list) > 0))

        # ---- Texte chunk
        text = f"Vraag: {vraag}\nAntwoord: {antw}"

        # ---- Métadonnées scalaires pour Chroma (CSV pour listes)
        gens_csv = ",".join(gens_list) if gens_list else ""
        tags_csv = ",".join(tags_list) if tags_list else ""
        title = (vraag[:80] + "…") if len(vraag) > 80 else vraag

        metas.append({
            "source": path,
            "title": title,
            "source_type": source_type,
            "categorie": category or "",
            "gens": gens_csv,              # CSV (pas de liste)
            "video_url": video or "",
            "photo": photo or "",
            "tags": tags_csv,              # CSV
            "sheet": chosen_sheet or "",
            "ask_gen": bool(ask_gen),      # bool OK pour Chroma
            "followup_q": followup_q or "",# petite phrase affichable côté /chat
        })
        texts.append(text)

        # ---- Index JSON (on garde les LISTES et les flags)
        index_rows.append({
            "question": vraag,
            "answer": antw,
            "category": category,
            "gens": gens_list,
            "video_url": video or None,
            "photo": photo or None,
            "tags": tags_list,
            "ask_gen": bool(ask_gen),
            "followup_q": followup_q or None,
            "source": path,
            "sheet": chosen_sheet,
        })

    if not texts:
        return {"indexed_files": 0, "indexed_chunks": 0, "error": "no rows"}

    # 6) Sauvegardes
    os.makedirs(STORE_DIR, exist_ok=True)
    index_path = os.path.join(STORE_DIR, "faq_index.json")
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index_rows, f, ensure_ascii=False, indent=2)

    vs = _get_vs()
    vs.add_texts(texts=texts, metadatas=metas)
    vs.persist()

    return {
        "indexed_files": 1,
        "indexed_chunks": len(texts),
        "sheet_used": chosen_sheet,
        "wrote_index": index_path,
    }

# ---------- Ingestion universelle (fichier ou dossier) ----------
def ingest_path(path: str, source_type: str = "mixed"):
    if not path:
        return {"indexed_files": 0, "indexed_chunks": 0}

    if os.path.isdir(path):
        return ingest_folder(path, source_type)

    ext = path.lower().rsplit(".", 1)[-1]
    if ext in {"xlsx", "xls"}:
        return ingest_excel(path, source_type="faq")

    texts, metas = [], []
    title = os.path.splitext(os.path.basename(path))[0]
    if ext == "pdf":
        for page_text, page_no in _load_pdf_pages(path):
            if not page_text.strip():
                continue
            chunks = [c for c in splitter.split_text(page_text) if c.strip()]
            texts.extend(chunks)
            metas.extend([{"source": path, "title": title, "source_type": source_type, "page": page_no}] * len(chunks))
    elif ext == "docx":
        text = _load_docx(path)
        chunks = [c for c in splitter.split_text(text) if c.strip()]
        texts.extend(chunks)
        metas.extend([{"source": path, "title": title, "source_type": source_type}] * len(chunks))
    elif ext in {"html", "htm"}:
        text = _load_html(path)
        chunks = [c for c in splitter.split_text(text) if c.strip()]
        texts.extend(chunks)
        metas.extend([{"source": path, "title": title, "source_type": source_type}] * len(chunks))
    elif ext in {"txt", "md"}:
        text = _load_txt(path)
        chunks = [c for c in splitter.split_text(text) if c.strip()]
        texts.extend(chunks)
        metas.extend([{"source": path, "title": title, "source_type": source_type}] * len(chunks))

    if not texts:
        return {"indexed_files": 0, "indexed_chunks": 0}

    vs = _get_vs()
    vs.add_texts(texts=texts, metadatas=metas)
    vs.persist()
    return {"indexed_files": 1, "indexed_chunks": len(texts)}
