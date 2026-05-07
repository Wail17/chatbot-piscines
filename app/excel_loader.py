# app/excel_loader.py
"""
Excel loader: AI 2.0.xlsx -> FAQAI.jsonl + extracted images.

Reads the owner's master Excel (AI 2.0.xlsx) and regenerates the
JSONL knowledge base with:
  - 336 entries (vs 166 in legacy JSONL)
  - Pre-translated EN/FR/DE columns stored as fields
  - image_path per row (PNG extracted from xl/media/ + drawings/)
  - video_url per row (Filmpje column)
  - Alternative phrasings (Alternatieve vragen column)
  - Dynamic synonym groups (synonims NL, Synonims EN sheets)
"""

from __future__ import annotations

import json
import logging
import os
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import openpyxl

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "AI 2.0.xlsx"
JSONL_OUT = PROJECT_ROOT / "app" / "data" / "all" / "faq" / "FAQAI.jsonl"
IMAGES_DIR = PROJECT_ROOT / "app" / "data" / "faq_images"

MAIN_SHEET = "Alle vragen"
SYN_NL_SHEET = "synonims NL"
SYN_EN_SHEET = "Synonims EN"

COL = {
    "id": 0,
    "category": 1,
    "vraag": 2,
    "antwoord": 3,
    "foto": 4,
    "filmpje": 5,
    "alternatieve": 6,
    "gen1": 7,
    "gen2": 8,
    "gen3": 9,
    "wifipool_alg": 10,
    "display": 11,
    "vloeibare_chloor": 12,
    "zoutelektrolyse": 13,
    "ph": 14,
    "redox": 15,
    "temperatuur": 16,
    "epdm": 17,
    "aut_kranen": 18,
    "frequentieregelaar": 19,
    "algemeen": 20,
    "hoe_installeren": 21,
    "onderhoud": 22,
    "de_frage": 23,
    "de_antwort": 24,
    "fr_question": 25,
    "fr_reponse": 26,
    "en_question": 27,
    "en_answer": 28,
}

XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XR_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _cell(row: Tuple[Any, ...], key: str) -> str:
    idx = COL.get(key)
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return "" if val is None else str(val).strip()


def _has_flag(row: Tuple[Any, ...], key: str) -> bool:
    v = _cell(row, key).lower()
    return v in {"x", "1", "true", "ja", "yes"}


def _extract_tags(row: Tuple[Any, ...]) -> List[str]:
    tags: List[str] = []
    flag_to_tag = {
        "gen1": "Gen1",
        "gen2": "Gen2",
        "gen3": "Gen3",
        "wifipool_alg": "Wifipool",
        "display": "Display",
        "vloeibare_chloor": "VloeibareChloor",
        "zoutelektrolyse": "Zoutelektrolyse",
        "ph": "pH",
        "redox": "Redox",
        "temperatuur": "Temperatuur",
        "epdm": "EPDM",
        "aut_kranen": "AutKranen",
        "frequentieregelaar": "Frequentieregelaar",
        "algemeen": "Algemeen",
        "hoe_installeren": "Installatie",
        "onderhoud": "Onderhoud",
    }
    for key, tag in flag_to_tag.items():
        if _has_flag(row, key):
            tags.append(tag)
    return tags


def _split_alt_questions(raw: str) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[\n\r;|]+", raw)
    return [p.strip() for p in parts if p.strip()]


def _clean_multiline(text: str) -> str:
    if not text:
        return ""
    return "\n".join(line.rstrip() for line in text.splitlines()).strip()


# Known first-word typos in the source Excel (missing first letter, etc.)
# Maps exact first word (case-sensitive) → correct first word.
FIRST_WORD_FIXES: Dict[str, str] = {
    "oer": "Voer",
}

# First words that ARE valid even when lowercase (technical terms, abbreviations).
LOWERCASE_ALLOWED_FIRST_WORDS = {"ph", "mv", "ms/cm", "rx", "orp"}


def _repair_answer(text: str) -> str:
    """Fix obvious defects in a FAQ answer pulled from the owner's Excel.

    Applies deterministic repairs only:
      1) Known first-word typos (e.g. 'oer' → 'Voer')
      2) Auto-capitalize the first letter when lowercase and the word isn't a
         technical abbreviation (pH, mV, RX, …) or a URL.
    Does NOT rewrite, rephrase, or translate.
    """
    if not text:
        return text
    stripped = text.lstrip()
    if not stripped:
        return text
    leading_ws = text[: len(text) - len(stripped)]

    first_word = stripped.split(None, 1)[0]
    remainder = stripped[len(first_word):]

    if first_word in FIRST_WORD_FIXES:
        return leading_ws + FIRST_WORD_FIXES[first_word] + remainder

    if first_word[:1].islower() and first_word.lower() not in LOWERCASE_ALLOWED_FIRST_WORDS:
        if not first_word.startswith(("http://", "https://", "www.")):
            return leading_ws + first_word[0].upper() + first_word[1:] + remainder

    return text


def _parse_drawing_row_map(xlsx_path: Path) -> Dict[str, Dict[int, int]]:
    """
    Parse xl/drawings/*.xml to map each drawing's embed rId -> row number.
    Returns: {drawing_file_name: {row_number_0_indexed: rId_int}}
    Only for the main "Alle vragen" sheet (drawing1.xml typically).
    """
    drawing_rows: Dict[str, Dict[str, int]] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            for name in z.namelist():
                if not (name.startswith("xl/drawings/") and name.endswith(".xml")):
                    continue
                content = z.read(name).decode("utf-8", errors="ignore")
                try:
                    root = ET.fromstring(content)
                except ET.ParseError:
                    continue
                row_to_rid: Dict[int, str] = {}
                for anchor in root.iter():
                    if not anchor.tag.endswith("twoCellAnchor") and not anchor.tag.endswith("oneCellAnchor"):
                        continue
                    from_el = anchor.find(f"{{{XDR_NS}}}from")
                    if from_el is None:
                        continue
                    row_el = from_el.find(f"{{{XDR_NS}}}row")
                    if row_el is None or not row_el.text:
                        continue
                    try:
                        row_num = int(row_el.text)
                    except ValueError:
                        continue
                    blip = None
                    for b in anchor.iter():
                        if b.tag.endswith("blip"):
                            blip = b
                            break
                    if blip is None:
                        continue
                    embed = blip.get(f"{{{REL_NS}}}embed")
                    if embed:
                        row_to_rid[row_num] = embed
                drawing_rows[name] = row_to_rid
    except Exception as e:
        logger.warning("Could not parse drawings: %s", e)
    return drawing_rows


def _parse_drawing_rels(xlsx_path: Path) -> Dict[str, Dict[str, str]]:
    """Parse xl/drawings/_rels/drawingN.xml.rels → rId → media target filename."""
    result: Dict[str, Dict[str, str]] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            for name in z.namelist():
                if not name.startswith("xl/drawings/_rels/") or not name.endswith(".rels"):
                    continue
                content = z.read(name).decode("utf-8", errors="ignore")
                try:
                    root = ET.fromstring(content)
                except ET.ParseError:
                    continue
                rid_to_target: Dict[str, str] = {}
                for rel in root.iter():
                    if not rel.tag.endswith("Relationship"):
                        continue
                    rid = rel.get("Id")
                    target = rel.get("Target") or ""
                    if rid and target:
                        clean = target.replace("../", "")
                        rid_to_target[rid] = clean
                drawing_base = os.path.basename(name).replace(".rels", "")
                result[f"xl/drawings/{drawing_base}"] = rid_to_target
    except Exception as e:
        logger.warning("Could not parse drawing rels: %s", e)
    return result


def _parse_sheet_drawing_link(xlsx_path: Path) -> Dict[str, str]:
    """Map sheet file (e.g., xl/worksheets/sheet1.xml) → drawing file path."""
    result: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            for name in z.namelist():
                if not name.startswith("xl/worksheets/_rels/") or not name.endswith(".rels"):
                    continue
                content = z.read(name).decode("utf-8", errors="ignore")
                try:
                    root = ET.fromstring(content)
                except ET.ParseError:
                    continue
                for rel in root.iter():
                    if not rel.tag.endswith("Relationship"):
                        continue
                    t = rel.get("Type") or ""
                    target = rel.get("Target") or ""
                    if t.endswith("/drawing") and target:
                        sheet_base = os.path.basename(name).replace(".rels", "")
                        sheet_path = f"xl/worksheets/{sheet_base}"
                        drawing_clean = target.replace("../", "xl/")
                        result[sheet_path] = drawing_clean
    except Exception as e:
        logger.warning("Could not parse sheet rels: %s", e)
    return result


def _resolve_sheet_file(xlsx_path: Path, sheet_name: str) -> Optional[str]:
    """Find xl/worksheets/sheetN.xml for a given sheet display name."""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="ignore")
            wb_rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="ignore")
        wb_root = ET.fromstring(wb_xml)
        rels_root = ET.fromstring(wb_rels)
        rid_for_sheet: Optional[str] = None
        for s in wb_root.iter():
            if s.tag.endswith("sheet") and s.get("name") == sheet_name:
                rid_for_sheet = s.get(f"{{{REL_NS}}}id")
                break
        if not rid_for_sheet:
            return None
        for rel in rels_root.iter():
            if rel.tag.endswith("Relationship") and rel.get("Id") == rid_for_sheet:
                target = rel.get("Target") or ""
                return f"xl/{target.replace('../', '').lstrip('/')}".replace("xl/xl/", "xl/")
    except Exception as e:
        logger.warning("Could not resolve sheet file: %s", e)
    return None


def extract_images(xlsx_path: Path, images_dir: Path) -> Dict[int, str]:
    """
    Extract images embedded in the main sheet and map them to row numbers (1-indexed, matching Excel row numbers).
    Saves to images_dir as row_<N>.<ext> and returns {excel_row: filename}.
    """
    images_dir.mkdir(parents=True, exist_ok=True)
    for old in images_dir.glob("row_*.*"):
        try:
            old.unlink()
        except OSError:
            pass

    sheet_file = _resolve_sheet_file(xlsx_path, MAIN_SHEET)
    sheet_to_drawing = _parse_sheet_drawing_link(xlsx_path)
    drawing_file = sheet_to_drawing.get(sheet_file or "")
    if not drawing_file:
        logger.warning("No drawing file linked to main sheet")
        return {}

    drawing_rows = _parse_drawing_row_map(xlsx_path).get(drawing_file, {})
    drawing_rels = _parse_drawing_rels(xlsx_path).get(drawing_file, {})

    row_to_filename: Dict[int, str] = {}
    with zipfile.ZipFile(xlsx_path) as z:
        for row_num, rid in drawing_rows.items():
            media_rel = drawing_rels.get(rid)
            if not media_rel:
                continue
            media_path = media_rel if media_rel.startswith("xl/") else f"xl/{media_rel}"
            try:
                data = z.read(media_path)
            except KeyError:
                continue
            ext = os.path.splitext(media_path)[1].lower() or ".png"
            excel_row = row_num + 1
            filename = f"row_{excel_row}{ext}"
            out_path = images_dir / filename
            with open(out_path, "wb") as f:
                f.write(data)
            row_to_filename[excel_row] = filename

    logger.info("Extracted %d images to %s", len(row_to_filename), images_dir)
    return row_to_filename


def load_synonym_sheets(xlsx_path: Path) -> List[List[str]]:
    """Load synonim sheets and return list of synonym groups (each a list of terms)."""
    groups: List[List[str]] = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        logger.warning("Could not open workbook for synonyms: %s", e)
        return []

    for sheet_name in (SYN_NL_SHEET, SYN_EN_SHEET):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            terms = [str(c).strip().lower() for c in row if c and str(c).strip()]
            if len(terms) >= 2:
                groups.append(terms)
    logger.info("Loaded %d synonym groups from Excel", len(groups))
    return groups


def build_entries(xlsx_path: Path, image_map: Dict[int, str]) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if MAIN_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{MAIN_SHEET}' not found in {xlsx_path}")
    ws = wb[MAIN_SHEET]

    entries: List[Dict[str, Any]] = []
    excel_row = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        excel_row += 1
        vraag = _cell(row, "vraag")
        antwoord = _cell(row, "antwoord")
        # If NL answer is empty, fall back to EN/FR/DE (some rows are only
        # documented in translated columns, e.g. row 132 Algemeen).
        if not antwoord:
            for fallback_key in ("en_answer", "fr_reponse", "de_antwort"):
                alt = _cell(row, fallback_key)
                if alt:
                    antwoord = alt
                    break
        if not antwoord:
            continue
        # Some rows (e.g. 27 "Wifipool Gen 1 reset", 28 "Wifipool gen 2 Reset")
        # have no explicit question — only a category/topic in column B. Fall
        # back to that topic, then to any translated question.
        if not vraag:
            for fallback_key in ("category", "en_question", "fr_question", "de_frage"):
                alt = _cell(row, fallback_key)
                if alt:
                    vraag = alt
                    break
            if not vraag:
                continue

        alt_raw = _cell(row, "alternatieve")
        alt_list = _split_alt_questions(alt_raw)

        vraag_clean = _clean_multiline(vraag)
        if "\n" in vraag_clean:
            parts = [p.strip() for p in vraag_clean.split("\n") if p.strip()]
            if len(parts) > 1:
                for extra in parts[1:]:
                    if extra not in alt_list:
                        alt_list.append(extra)
                vraag_clean = parts[0]

        entry: Dict[str, Any] = {
            "Categorie": _cell(row, "category"),
            "Vraag": vraag_clean,
            "Antwoord": _repair_answer(_clean_multiline(antwoord)),
            "alt_questions": alt_list,
            "tags": _extract_tags(row),
            "excel_row": excel_row,
        }

        filmpje = _cell(row, "filmpje")
        if filmpje and (filmpje.startswith("http") or "youtube" in filmpje.lower() or "youtu.be" in filmpje.lower()):
            entry["video_url"] = filmpje

        if excel_row in image_map:
            rel = f"/faq_images/{image_map[excel_row]}"
            entry["image_path"] = rel

        de_q = _cell(row, "de_frage")
        de_a = _cell(row, "de_antwort")
        if de_q:
            entry["DEFrage"] = de_q
        if de_a:
            entry["DEAntwort"] = _repair_answer(_clean_multiline(de_a))

        fr_q = _cell(row, "fr_question")
        fr_a = _cell(row, "fr_reponse")
        if fr_q:
            entry["FRQuestion"] = fr_q
        if fr_a:
            entry["FRReponse"] = _repair_answer(_clean_multiline(fr_a))

        en_q = _cell(row, "en_question")
        en_a = _cell(row, "en_answer")
        if en_q:
            entry["ENQuestion"] = en_q
        if en_a:
            entry["ENAnswer"] = _repair_answer(_clean_multiline(en_a))

        entries.append(entry)

    logger.info("Built %d entries from Excel", len(entries))
    return entries


def _llm_polish(text: str, lang_name: str) -> Optional[str]:
    """Send one answer through an LLM to clean typos, missing letters, grammar.

    Priorities, in order: Anthropic (Haiku), OpenAI (gpt-4o-mini).
    Returns cleaned text, or None if no LLM is available / call failed.
    The LLM is instructed to fix ONLY defects — never rewrite or add content.
    """
    if not text or not text.strip():
        return None

    prompt = (
        f"You are a careful proofreader for a {lang_name} FAQ knowledge base. "
        f"Fix defects in the answer below:\n"
        f"- missing letters at the start of words (e.g. 'oer' → 'Voer')\n"
        f"- obvious typos and misspellings (e.g. 'collecotor' → 'collector', 'overbang' → 'overgang')\n"
        f"- capitalization (sentences must start with a capital letter)\n"
        f"- missing punctuation at the end of sentences\n"
        f"- broken spacing\n\n"
        f"STRICT RULES — never violate:\n"
        f"- Do NOT rewrite, rephrase, translate, shorten, or expand.\n"
        f"- Keep the content, facts, order, and meaning identical.\n"
        f"- Keep line breaks, bullet points, numbers, URLs, product names, and technical terms exactly as written.\n"
        f"- If the text is already correct, return it UNCHANGED.\n"
        f"- Reply with ONLY the corrected text — no preamble, no explanation, no quotes.\n\n"
        f"Answer:\n{text}"
    )

    try:
        from .rag import anthropic_client, openai_client
        from .config import LLM_MODEL
    except Exception:
        return None

    if anthropic_client is not None:
        try:
            resp = anthropic_client.messages.create(
                model=LLM_MODEL,
                max_tokens=2048,
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}],
            )
            out = (resp.content[0].text or "").strip()
            if out and len(out) >= max(10, int(len(text) * 0.5)):
                return out
        except Exception as e:
            logger.debug("Anthropic polish failed: %s", e)

    if openai_client is not None:
        try:
            resp = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                temperature=0.0,
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )
            out = (resp.choices[0].message.content or "").strip()
            if out and len(out) >= max(10, int(len(text) * 0.5)):
                return out
        except Exception as e:
            logger.debug("OpenAI polish failed: %s", e)

    return None


def polish_entries_with_llm(entries: List[Dict[str, Any]], max_workers: int = 8) -> int:
    """In-place LLM polish pass over every answer field in every entry.

    Polished text overwrites the original only when the LLM returns something
    non-empty and not trivially shorter. Returns the count of rewrites.
    Skips silently when no LLM is available so reloads still succeed offline.

    Runs in parallel (threads) since Anthropic API calls are I/O bound.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    field_lang = {
        "Antwoord": "Dutch",
        "ENAnswer": "English",
        "FRReponse": "French",
        "DEAntwort": "German",
    }

    jobs = []
    for entry in entries:
        for field, lang_name in field_lang.items():
            orig = entry.get(field)
            if not orig or not isinstance(orig, str):
                continue
            jobs.append((entry, field, lang_name, orig))

    total = len(jobs)
    logger.info("LLM polish: %d fields to process (workers=%d)", total, max_workers)
    if total == 0:
        return 0

    updated = 0
    done = 0

    def work(item):
        entry, field, lang_name, orig = item
        polished = _llm_polish(orig, lang_name)
        return entry, field, orig, polished

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [pool.submit(work, j) for j in jobs]
        for fut in as_completed(futures):
            done += 1
            try:
                entry, field, orig, polished = fut.result()
            except Exception as e:
                logger.debug("polish worker error: %s", e)
                continue
            if polished and polished.strip() and polished.strip() != orig.strip():
                entry[field] = polished
                updated += 1
            if done % 50 == 0 or done == total:
                logger.info("LLM polish progress: %d/%d (%d rewrites so far)", done, total, updated)

    return updated


def write_jsonl(entries: List[Dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        for entry in entries:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    os.replace(tmp_path, out_path)
    logger.info("Wrote %d entries to %s", len(entries), out_path)


def write_synonyms_extra(groups: List[List[str]], out_path: Path) -> None:
    """Persist Excel synonym groups to a JSON file for the synonyms module to load."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(groups, f, ensure_ascii=False, indent=2)
    logger.info("Wrote %d synonym groups to %s", len(groups), out_path)


def reload_from_excel(
    excel_path: Path = EXCEL_PATH,
    jsonl_out: Path = JSONL_OUT,
    images_dir: Path = IMAGES_DIR,
    polish: bool = False,
) -> Dict[str, int]:
    """Full reload pipeline. Returns summary dict.

    When `polish=True`, every answer is sent through an LLM (Anthropic or
    OpenAI) to fix typos, capitalization, missing letters, and grammar —
    without rewriting the content. This adds ~1 LLM call per answer (runs
    once per reload, then served from JSONL).
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    image_map = extract_images(excel_path, images_dir)
    entries = build_entries(excel_path, image_map)

    polished_count = 0
    if polish:
        logger.info("Running LLM polish pass over %d entries...", len(entries))
        polished_count = polish_entries_with_llm(entries)
        logger.info("LLM polish done: %d answers rewritten", polished_count)

    write_jsonl(entries, jsonl_out)

    syn_groups = load_synonym_sheets(excel_path)
    syn_out = PROJECT_ROOT / "app" / "data" / "excel_synonyms.json"
    write_synonyms_extra(syn_groups, syn_out)

    return {
        "entries": len(entries),
        "images": len(image_map),
        "synonym_groups": len(syn_groups),
        "polished": polished_count,
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Reload FAQ from AI 2.0.xlsx")
    parser.add_argument("--reload", action="store_true", help="Perform full reload")
    parser.add_argument("--polish", action="store_true",
                        help="Run an LLM polish pass to fix typos/grammar in every answer")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass

    summary = reload_from_excel(polish=args.polish)
    print(f"Entries:  {summary['entries']}")
    print(f"Images:   {summary['images']}")
    print(f"Synonym groups (Excel): {summary['synonym_groups']}")
    print(f"Polished: {summary.get('polished', 0)}")
